# -- coding: utf-8 --
import time
start_time = time.time()

import openpyxl
import random

# 新建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# 写入表头
ws.cell(row=1, column=1).value = "姓名"
ws.cell(row=1, column=2).value = "学号"
ws.cell(row=1, column=3).value = "成绩"

# 随机生成50个人的数据
for i in range(2, 52):
    # 随机生成姓名
    first_name = random.choice(["张", "李", "王", "赵", "刘", "陈", "杨", "黄", "吴", "周"])
    last_name = random.choice(["三", "四", "五", "六", "七", "八", "九", "十"])
    name = first_name + last_name
    # 随机生成学号
    student_id = "20" + str(random.randint(100000, 999999))
    # 随机生成成绩
    score = random.randint(71, 89)
    # 写入数据
    ws.cell(row=i, column=1).value = name
    ws.cell(row=i, column=2).value = student_id
    ws.cell(row=i, column=3).value = score

# 保存文件
wb.save("1随机生成.xlsx")
print("文件生成成功！")

end_time = time.time()
total_time = end_time - start_time
print("\033[31m程序运行时间为：{:.2f}秒\033[0m".format(total_time))

